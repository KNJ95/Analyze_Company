"""
export_excel.py
public/jobs_raw.json を 新卒求人収集テンプレート.xlsx の形式で出力する。

使い方:
  python export_excel.py                          # jobs_raw.json → 求人データ.xlsx
  python export_excel.py --template template.xlsx # テンプレートを指定
  python export_excel.py --out output.xlsx        # 出力先を指定
"""

import json, argparse, pathlib
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 列定義（テンプレートの列順に対応） ────────────────────────────────────────
COLUMNS = [
    ("会社名",                "company"),
    ("業種",                  "industry"),
    ("企業規模",               "size_label"),
    ("売上（任意）",            None),
    ("職種",                  "job_types"),       # リストを結合
    ("勤務地",                 None),
    ("仕事内容（全文）",         "outline_summary"),
    ("求める人物像（最重要）",    "persona"),
    ("必須スキル",              None),
    ("歓迎スキル",              None),
    ("求める行動特性",           "traits"),         # リストを結合
    ("使用技術・ツール",          None),
    ("初任給",                 None),
    ("URL（求人ページ）",        "url"),
    ("取得日",                 "scraped_at"),
    ("メモ",                  None),
    # Be-Ready 評価軸（テンプレートに追加）
    ("Be-ReadyレベルLv1〜4",  "level"),
    ("①課題設定力",            "s1"),
    ("②情報活用力",            "s2"),
    ("③不確実性耐性",           "s3"),
    ("④提案・発信力",           "s4"),
    ("⑤実行・改善力",           "s5"),
    ("⑥オーナーシップ",         "s6"),
    ("⑦協働・調整力",           "s7"),
    ("⑧自律・内発的動機",        "s8"),
    ("⑨行動変容力",            "s9"),
    ("平均スコア",              "avg_score"),
    ("判定理由",               "level_reason"),
]

# ─── カラー ────────────────────────────────────────────────────────────────────
PURPLE_DARK  = "460073"
PURPLE_MID   = "7500C0"
PURPLE_LIGHT = "A100FF"
PURPLE_BG    = "F3E8FF"
WHITE        = "FFFFFF"
GRAY         = "F5F5F5"

SCORE_COLORS = {
    5: ("460073", "FFFFFF"),
    4: ("7500C0", "FFFFFF"),
    3: ("A100FF", "FFFFFF"),
    2: ("D4AAFF", "2D0060"),
    1: ("E8E8E8", "666666"),
    0: ("FFFFFF", "AAAAAA"),
}

def score_col(n):
    """スコア列かどうか判定（s1〜s9）"""
    return n in [f"s{i}" for i in range(1, 10)]

def cell_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def cell_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def thin_border():
    s = Side(border_style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

# ─── データ整形 ────────────────────────────────────────────────────────────────

def extract_value(job, key):
    """jobs_raw.json の1件から列キーに対応する値を取得する"""
    if key is None:
        return ""
    scores = (job.get("result") or {}).get("scores", {})
    result = job.get("result") or {}

    if key == "company":
        return result.get("company") or job.get("corpId", "")
    if key == "industry":
        return result.get("industry", "")
    if key == "size_label":
        return result.get("size_label", "")
    if key == "job_types":
        return " / ".join(result.get("job_types", []))
    if key == "outline_summary":
        # outlineText の先頭300文字を要約として使用
        text = job.get("outlineText", "")
        return text[:300].replace("\n", " ").strip() if text else ""
    if key == "persona":
        return result.get("persona", "")
    if key == "traits":
        return " / ".join(result.get("traits", []))
    if key == "url":
        corp_id = job.get("corpId", "")
        return f"https://job.mynavi.jp/27/pc/search/corp{corp_id}/outline.html" if corp_id else ""
    if key == "scraped_at":
        return job.get("scraped_at") or str(date.today())
    if key == "level":
        return result.get("level", "")
    if key == "level_reason":
        return result.get("level_reason", "")
    if key == "avg_score":
        vals = [v for v in scores.values() if isinstance(v, (int, float)) and v > 0]
        return round(sum(vals) / len(vals), 1) if vals else ""
    if key.startswith("s") and key[1:].isdigit():
        axis_id = key[1:]
        v = scores.get(str(axis_id)) or scores.get(int(axis_id))
        return int(v) if v else 0
    return ""

# ─── Excel出力 ─────────────────────────────────────────────────────────────────

def export(jobs_path: str, output_path: str):
    # データ読み込み
    raw = json.loads(pathlib.Path(jobs_path).read_text(encoding="utf-8"))
    # 分析済みのみ（resultがある件）
    done = [j for j in raw if j.get("result")]
    print(f"分析済み: {len(done)}件 / 全体: {len(raw)}件")

    wb = Workbook()
    ws = wb.active
    ws.title = "新卒求人データ"

    # ── ヘッダー行 ──────────────────────────────────────────────────────────
    for col_idx, (header, key) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        # Be-Ready軸列は紫
        if key and (key.startswith("s") and key[1:].isdigit() or key in ("level","avg_score","level_reason")):
            cell.fill = cell_fill(PURPLE_DARK)
            cell.font = cell_font(bold=True, color=WHITE, size=10)
        else:
            cell.fill = cell_fill(PURPLE_MID)
            cell.font = cell_font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[1].height = 32

    # ── データ行 ──────────────────────────────────────────────────────────
    for row_idx, job in enumerate(done, start=2):
        bg = GRAY if row_idx % 2 == 0 else WHITE
        for col_idx, (header, key) in enumerate(COLUMNS, start=1):
            value = extract_value(job, key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()

            # スコア列：カラーヒートマップ
            if key and key.startswith("s") and key[1:].isdigit() and isinstance(value, int):
                sc = value if 1 <= value <= 5 else 0
                fill_hex, font_hex = SCORE_COLORS.get(sc, SCORE_COLORS[0])
                cell.fill = cell_fill(fill_hex)
                cell.font = cell_font(bold=(sc >= 4), color=font_hex, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif key == "avg_score" and value:
                cell.fill = cell_fill(PURPLE_BG)
                cell.font = cell_font(bold=True, color=PURPLE_DARK, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0.0"
            elif key == "level":
                cell.fill = cell_fill(PURPLE_BG)
                cell.font = cell_font(bold=True, color=PURPLE_DARK)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif key == "url":
                cell.font = cell_font(color="0070C0", size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.fill = cell_fill(bg)
                cell.font = cell_font(size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws.row_dimensions[row_idx].height = 30

    # ── 列幅設定 ────────────────────────────────────────────────────────
    col_widths = {
        1:  20,   # 会社名
        2:  14,   # 業種
        3:  12,   # 規模
        4:  10,   # 売上
        5:  18,   # 職種
        6:  12,   # 勤務地
        7:  40,   # 仕事内容
        8:  40,   # 求める人物像
        9:  20,   # 必須スキル
        10: 20,   # 歓迎スキル
        11: 25,   # 行動特性
        12: 18,   # 技術
        13: 12,   # 初任給
        14: 35,   # URL
        15: 12,   # 取得日
        16: 16,   # メモ
        17: 14,   # Lv
        18: 8, 19: 8, 20: 8, 21: 8, 22: 8,  # 軸①〜⑤
        23: 8, 24: 8, 25: 8, 26: 8,           # 軸⑥〜⑨
        27: 10,   # 平均
        28: 30,   # 判定理由
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ウィンドウ固定（1行目 + A列）
    ws.freeze_panes = "B2"

    # ── 集計シート ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("業種別集計")
    _write_summary(ws2, done)

    wb.save(output_path)
    print(f"✓ 出力完了: {output_path}（{len(done)}件）")


def _write_summary(ws, jobs):
    """業種別の評価軸平均スコアを集計してシートに出力"""
    from collections import defaultdict

    summary = defaultdict(lambda: {"count": 0, "sums": {}, "cnts": {}})
    for j in jobs:
        ind = (j.get("result") or {}).get("industry") or "不明"
        summary[ind]["count"] += 1
        scores = (j.get("result") or {}).get("scores", {})
        for i in range(1, 10):
            s = scores.get(str(i)) or scores.get(i)
            if s and int(s) > 0:
                summary[ind]["sums"][i] = summary[ind]["sums"].get(i, 0) + int(s)
                summary[ind]["cnts"][i] = summary[ind]["cnts"].get(i, 0) + 1

    axis_names = ["課題設定力","情報活用力","不確実性耐性","提案・発信力","実行・改善力",
                  "オーナーシップ","協働・調整力","自律・内発的動機","行動変容力"]

    # タイトル
    ws.merge_cells("A1:L1")
    ws["A1"] = "業種別 Be-Ready評価軸 重要度マップ"
    ws["A1"].fill = cell_fill(PURPLE_DARK)
    ws["A1"].font = cell_font(bold=True, color=WHITE, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ヘッダー
    headers = ["業種", "件数"] + [f"軸{i}\n{n}" for i, n in enumerate(axis_names, 1)] + ["平均"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = cell_fill(PURPLE_MID)
        c.font = cell_font(bold=True, color=WHITE, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[2].height = 34

    # データ
    for ri, (ind, d) in enumerate(sorted(summary.items(), key=lambda x: -x[1]["count"]), start=3):
        ws.cell(row=ri, column=1, value=ind).fill = cell_fill(PURPLE_LIGHT)
        ws.cell(row=ri, column=1).font = cell_font(bold=True, color=WHITE)
        ws.cell(row=ri, column=1).border = thin_border()

        ws.cell(row=ri, column=2, value=d["count"]).fill = cell_fill(PURPLE_BG)
        ws.cell(row=ri, column=2).font = cell_font(bold=True, color=PURPLE_DARK)
        ws.cell(row=ri, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=ri, column=2).border = thin_border()

        axis_avgs = []
        for ci, i in enumerate(range(1, 10), start=3):
            s, c = d["sums"].get(i, 0), d["cnts"].get(i, 0)
            avg = round(s / c, 1) if c else 0
            if avg: axis_avgs.append(avg)
            sc = round(avg) if avg else 0
            fill_hex, font_hex = SCORE_COLORS.get(sc, SCORE_COLORS[0])
            cell = ws.cell(row=ri, column=ci, value=avg if avg else "—")
            cell.fill = cell_fill(fill_hex) if avg else cell_fill("F8F8F8")
            cell.font = cell_font(bold=(sc >= 4), color=font_hex if avg else "CCCCCC", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()

        total_avg = round(sum(axis_avgs) / len(axis_avgs), 1) if axis_avgs else ""
        c_avg = ws.cell(row=ri, column=12, value=total_avg)
        c_avg.fill = cell_fill(PURPLE_BG)
        c_avg.font = cell_font(bold=True, color=PURPLE_DARK, size=11)
        c_avg.alignment = Alignment(horizontal="center", vertical="center")
        c_avg.border = thin_border()
        ws.row_dimensions[ri].height = 26

    # 列幅
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 8
    for i in range(3, 13):
        ws.column_dimensions[get_column_letter(i)].width = 10


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="jobs_raw.json → Excel出力")
    parser.add_argument("--input",  default="public/jobs_raw.json", help="入力JSONファイル")
    parser.add_argument("--out",    default="求人分析結果.xlsx",      help="出力Excelファイル名")
    args = parser.parse_args()

    export(args.input, args.out)